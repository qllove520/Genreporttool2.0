import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QFileDialog, QMessageBox
)
from PyQt5.QtCore import Qt
from core.excel_utils import find_row_by_fuzzy_column_value,write_to_target_sheet
from openpyxl import load_workbook

class ExcelTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("项目台账写入工具")
        self.setMinimumSize(700, 500)

        self.data_file = ""
        self.template_file = ""

        # 新增输入字段变量
        self.input_fields = {}

        self.init_ui()

    def init_ui(self):
        n=0
        layout = QVBoxLayout()

        # 数据台账
        self.label_data = QLabel("未选择 数据台账 文件")
        btn_data = QPushButton("选择 数据台账（读取源）")
        btn_data.clicked.connect(self.choose_data_file)

        # 写入模板
        self.label_template = QLabel("未选择 写入模板 文件")
        btn_template = QPushButton("选择 写入模板（目标文件）")
        btn_template.clicked.connect(self.choose_template_file)

        layout.addWidget(self.label_data)
        layout.addWidget(btn_data)
        layout.addWidget(self.label_template)
        layout.addWidget(btn_template)

        # 关键词
        layout.addWidget(QLabel("关键词（如：2600E2）："))
        self.input_keyword = QLineEdit()
        layout.addWidget(self.input_keyword)

        # 额外字段
        extra_fields = [
            "测试单号", "申请理由", "开始时间", "结束时间", "测试依据", "测试范围"
        ]

        extra_tips =["CPKFLX20240426001","新产品导入","2025/07/14","2025/07/21","窗口式照相机技术需求规格书V1.3.xlsx","回归上一轮Bug、按照整机验收标准：底层软件、上层应用、机电安全"]

        for field in extra_fields:
            row = QHBoxLayout()
            label = QLabel(field + ":")
            input_box = QLineEdit()
            input_box.setPlaceholderText(f"可选填写 例如： {extra_tips[n]}")
            row.addWidget(label)
            row.addWidget(input_box)
            layout.addLayout(row)
            self.input_fields[field] = input_box
            n+=1

        # 提交按钮
        btn_process = QPushButton("提取并写入")
        btn_process.clicked.connect(self.process)
        layout.addWidget(btn_process)

        self.setLayout(layout)

    def choose_data_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 数据台账 文件", "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.data_file = path
            self.label_data.setText(f"已选择数据台账: {path}")

    def choose_template_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择 写入模板 文件", "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.template_file = path
            self.label_template.setText(f"已选择模板文件: {path}")

    def process(self):
        if not self.data_file or not self.template_file:
            QMessageBox.warning(self, "错误", "请先选择 数据台账 和 写入模板")
            return

        keyword = self.input_keyword.text().strip()
        if not keyword:
            QMessageBox.warning(self, "错误", "请输入关键词")
            return

        try:
            # 提取主数据
            target_fields = ['项目编号', '项目名称', '内部型号', '产品名称', '项目经理', '产品经理', '负责人']
            cell_mapping = {
                '项目编号': 'D2',
                '项目名称': 'H2',
                '项目经理': 'U2',
                '内部型号': 'D3',
                '产品名称': 'H3',
                '产品经理': 'U3',
                '负责人': 'U4'
            }

            result = find_row_by_fuzzy_column_value(
                file_path=self.data_file,
                key_column='项目_产品',
                key_value=keyword,
                target_columns=target_fields
            )

            if not result:
                QMessageBox.warning(self, "未找到", "台账中未找到匹配行")
                return

            # 提取额外输入内容
            extra_data = {
                field: self.input_fields[field].text().strip()
                for field in self.input_fields
                if self.input_fields[field].text().strip()
            }

            extra_cell_mapping = {
                '测试单号': 'O2',
                '申请理由': 'D4',
                '开始时间': 'H4',
                '结束时间': 'O4',
                '测试依据': 'E6',
                '测试范围': 'E7'
            }

            # 合并数据并写入
            wb = load_workbook(self.template_file,keep_vba=True)
            sheet_name = '验收测试结果'
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"写入模板缺少工作表：{sheet_name}")
            sheet = wb[sheet_name]

            # 主数据写入
            for key, cell in cell_mapping.items():
                sheet[cell] = result.get(key, "")

            # 附加字段写入（仅填写的才写）
            for key, cell in extra_cell_mapping.items():
                if key in extra_data:
                    sheet[cell] = extra_data[key]

            wb.save(self.template_file)

            QMessageBox.information(self, "成功", "数据已成功写入 Excel 模板！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"发生错误：\n{str(e)}")

