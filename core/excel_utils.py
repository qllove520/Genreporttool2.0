import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import xlwings as xw
import pandas as pd




def find_row_by_fuzzy_column_value(file_path, key_column, key_value, target_columns):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers)}

    if key_column not in header_index:
        raise ValueError(f"找不到列标题: {key_column}")
    for col in target_columns:
        if col not in header_index:
            raise ValueError(f"找不到目标列标题: {col}")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        cell_value = str(row[header_index[key_column]])
        if key_value in cell_value:
            return {col: row[header_index[col]] for col in target_columns}
    return None


def write_to_target_sheet(file_path, sheet_name, cell_map, data_dict):
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表：{sheet_name}")
    sheet = wb[sheet_name]
    for key, cell in cell_map.items():
        sheet[cell] = data_dict.get(key, "")
    wb.save(file_path)

def fill_excel_template_acceptance(template_path: str, data: dict, field_mapping: dict, sheet_name: str, log_callback=None):
    """
    Fills an Excel template with user input data, handles merged cells.
    Used for the acceptance test filling page.
    """
    if not os.path.exists(template_path):
        if log_callback: log_callback(f"错误: Excel 模板文件未找到于 '{template_path}'", is_error=True)
        return False
    if not template_path.lower().endswith((".xlsx", ".xlsm")):
        if log_callback: log_callback(f"错误: 提供的文件 '{template_path}' 不是有效的 Excel 模板 (.xlsx 或 .xlsm)。", is_error=True)
        return False

    try:
        wb = load_workbook(template_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if log_callback: log_callback(f"已成功加载工作表: '{sheet_name}'。")
        else:
            if log_callback: log_callback(f"错误: Excel 工作簿中未找到名为 '{sheet_name}' 的工作表。请检查工作表名称是否正确。", is_error=True)
            return False
    except Exception as e:
        if log_callback: log_callback(f"错误: 无法加载 Excel 工作簿或获取指定工作表 '{template_path}'。原因: {e}", is_error=True)
        return False

    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_coord = ws.cell(row=min_row, column=min_col).coordinate
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = ws.cell(row=row, column=col).coordinate
                merged_cells_map[coord] = top_left_coord

    output_file_name = "filled_" + os.path.basename(template_path)
    output_path = os.path.join(os.path.dirname(template_path), output_file_name)

    if log_callback: log_callback("正在写入数据到 Excel...")
    all_fields_processed_successfully = True

    for field_name, config in field_mapping.items():
        excel_cell_coord = config["excel_cell"]
        value = data.get(field_name, "")

        if excel_cell_coord:
            actual_cell_coord = merged_cells_map.get(excel_cell_coord, excel_cell_coord)
            try:
                if value:
                    ws[actual_cell_coord] = value
                    if log_callback: log_callback(f"  写入字段 '{field_name}': '{value}' 到单元格 '{actual_cell_coord}'")
                else:
                    if log_callback: log_callback(f"  跳过字段 '{field_name}': 未填写内容，单元格 '{actual_cell_coord}' 保持不变。", is_error=False)
            except Exception as e:
                if log_callback: log_callback(f"  写入字段 '{field_name}' 到单元格 '{actual_cell_coord}' 失败。原因: {e}", is_error=True)
                all_fields_processed_successfully = False
        else:
            if log_callback: log_callback(f"  警告: 字段 '{field_name}' 在配置中未指定 Excel 单元格，跳过写入。", is_error=True)
            all_fields_processed_successfully = False

    try:
        wb.save(output_path)
        if all_fields_processed_successfully:
            if log_callback: log_callback(f"\n--- 成功填充！文件保存为: '{output_path}' ---", is_error=False)
        else:
            if log_callback: log_callback(f"\n--- 填充完成，但有部分字段出现问题。文件保存为: '{output_path}' ---", is_error=False)
            if log_callback: log_callback("注意: 请检查日志，有部分字段未填写内容或写入失败。", is_error=True)
        return True
    except PermissionError:
        if log_callback: log_callback(f"错误: 无法保存文件 '{output_path}'。原因: 权限被拒绝，请确保 Excel 文件已关闭且您有写入权限。", is_error=True)
        return False
    except Exception as e:
        if log_callback: log_callback(f"错误: 无法将填充后的 Excel 文件保存到 '{output_path}'。原因: {e}", is_error=True)
        return False


def consolidate_excel_data_and_insert_chart(doc1_path: str, doc2_path: str, doc3_path: str, doc4_path: str,
                                            target_report_path: str, log_callback=None):
    """
    Core data consolidation logic using xlwings: copies three data tables (starting from the second row)
    to the third row of corresponding sheets in the target report, preserving format and sheet order.
    Also inserts the device appearance image into the '设备外观图' sheet.
    This version allows individual source documents (Doc1-Doc4) to be optional.
    """
    source_info = [
        {'path': doc1_path, 'sheet_name': '遗留缺陷列表'},
        {'path': doc2_path, 'sheet_name': '产品需求列表'},
        {'path': doc3_path, 'sheet_name': '验收测试用例'}
    ]

    app = None
    try:
        # Check target report path first as it's mandatory
        if not target_report_path or not os.path.exists(target_report_path):
            # FIXED: Always pass is_error
            if log_callback: log_callback(
                f"错误：目标报告文件 '{os.path.basename(target_report_path) if target_report_path else '未指定'}' 不存在或路径为空。",
                True)
            return False

        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(target_report_path, update_links=False)
        # FIXED: Always pass is_error
        if log_callback: log_callback(f"已打开目标报告：{os.path.basename(target_report_path)}", False)

        for info in source_info:
            src_path = info['path']
            target_sheet_name = info['sheet_name']

            if not src_path or not os.path.exists(src_path):
                # FIXED: Always pass is_error
                if log_callback: log_callback(
                    f"警告：源文件 '{os.path.basename(src_path) if src_path else target_sheet_name + '文档'}' 未选择或不存在，跳过处理。",
                    False)
                continue  # Skip to the next source file

            # FIXED: Always pass is_error
            if log_callback: log_callback(
                f"\n正在处理源文件 '{os.path.basename(src_path)}' -> 工作表 '{target_sheet_name}'", False)

            try:
                df = pd.read_excel(src_path, header=None)
                if df.empty:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"警告：源文件 '{os.path.basename(src_path)}' 为空或无数据。", False)
                    data = []
                else:
                    data = df.iloc[1:].values.tolist()  # Skip the first row (header in ZenTao exports)
            except Exception as e:
                # FIXED: Always pass is_error
                if log_callback: log_callback(f"错误: 读取源文件 '{os.path.basename(src_path)}' 失败。原因: {e}", True)
                if log_callback: log_callback(traceback.format_exc(), True)
                continue  # Skip to the next source file

            if target_sheet_name in [s.name for s in wb.sheets]:
                # FIXED: Always pass is_error
                sht = wb.sheets[target_sheet_name]
                if log_callback: log_callback(f"已找到工作表: '{target_sheet_name}'", False)
            else:
                try:
                    sht = wb.sheets.add(name=target_sheet_name, after=wb.sheets[-1])  # Add new sheet at the end
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"创建工作表：{target_sheet_name}", False)
                except Exception as e:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"错误: 无法创建工作表 '{target_sheet_name}'。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
                    # Don't return, continue to save if other operations were successful
                    sht = None  # Ensure sht is None if creation failed

            if sht:  # Only proceed if sheet exists or was created
                start_row_excel = 3
                used_range = sht.used_range
                last_row_to_clear = used_range.last_cell.row if not used_range.api is None else start_row_excel
                last_col_to_clear = used_range.last_cell.column if not used_range.api is None else (
                    df.shape[1] if not df.empty else 1)

                if data:
                    max_cols_in_data = max(len(row) for row in data)
                    last_col_to_clear = max(last_col_to_clear, max_cols_in_data)

                if last_row_to_clear >= start_row_excel:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(
                        f"清除 '{target_sheet_name}' 第 {start_row_excel} 行到第 {last_row_to_clear} 行的内容 (到第 {last_col_to_clear} 列)...",
                        False)
                    try:
                        sht.range((start_row_excel, 1), (last_row_to_clear, last_col_to_clear)).clear_contents()
                    except Exception as e:
                        # FIXED: Always pass is_error
                        if log_callback: log_callback(f"错误: 清除工作表 '{target_sheet_name}' 旧数据失败。原因: {e}",
                                                      True)
                        if log_callback: log_callback(traceback.format_exc(), True)
                        continue  # Try to proceed, but log error
                else:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"工作表 '{target_sheet_name}' 已经足够干净，无需清除旧数据。", False)

                if data:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"粘贴新数据到 '{target_sheet_name}'，共 {len(data)} 行。", False)
                    try:
                        sht.range((start_row_excel, 1)).value = data
                    except Exception as e:
                        # FIXED: Always pass is_error
                        if log_callback: log_callback(f"错误: 粘贴数据到工作表 '{target_sheet_name}' 失败。原因: {e}",
                                                      True)
                        if log_callback: log_callback(traceback.format_exc(), True)
                        continue
                else:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"没有数据需要粘贴到 '{target_sheet_name}'。", False)

        # Handle image insertion (Doc4)
        pic_sheet_name = '设备外观图'
        if doc4_path and os.path.exists(doc4_path):
            # FIXED: Always pass is_error
            if log_callback: log_callback(f"\n正在处理图片文件 '{os.path.basename(doc4_path)}'", False)
            if pic_sheet_name in [s.name for s in wb.sheets]:
                # FIXED: Always pass is_error
                pic_sht = wb.sheets[pic_sheet_name]
                if log_callback: log_callback(f"已找到工作表: '{pic_sheet_name}'", False)
            else:
                try:
                    pic_sht = wb.sheets.add(name=pic_sheet_name, after=wb.sheets[-1])
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"创建工作表: '{pic_sheet_name}'", False)
                except Exception as e:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"错误: 无法创建图片工作表 '{pic_sheet_name}'。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
                    # Don't return, continue to save if other operations were successful
                    pic_sht = None  # Ensure pic_sht is None if creation failed

            if pic_sht:  # Only proceed if sheet exists or was created
                try:
                    second_row_top = pic_sht.range('2:2').top if not pic_sht.used_range.api is None else float('inf')
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"正在清除 '{pic_sheet_name}' 中第二行及以后所有图片...", False)
                    pictures_deleted_count = 0
                    for pic in list(pic_sht.pictures):
                        if pic.top >= second_row_top:
                            pic.delete()
                            pictures_deleted_count += 1
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"已清除 {pictures_deleted_count} 张图片。", False)

                    width_pt = 23.66 * 28.3465
                    height_pt = 13.31 * 28.3465
                    top_left_cell = pic_sht.range('A2')

                    normalized_doc4_path = os.path.normpath(doc4_path)
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(
                        f"准备插入图片。原始路径: '{doc4_path}', 规范化路径: '{normalized_doc4_path}'", False)
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(
                        f"正在插入图片 '{os.path.basename(normalized_doc4_path)}' 到 '{pic_sheet_name}' 的 '{top_left_cell.address}'...",
                        False)
                    pic_sht.pictures.add(normalized_doc4_path,
                                         left=top_left_cell.left,
                                         top=top_left_cell.top,
                                         width=width_pt,
                                         height=height_pt)
                    # FIXED: Always pass is_error
                    if log_callback: log_callback("图片插入成功。", False)
                except Exception as e:
                    # FIXED: Always pass is_error
                    if log_callback: log_callback(f"错误: 插入图片到工作表 '{pic_sheet_name}' 失败。原因: {e}", True)
                    if log_callback: log_callback(traceback.format_exc(), True)
        else:
            # FIXED: Always pass is_error
            if log_callback: log_callback(
                f"警告：图片文件 '{os.path.basename(doc4_path) if doc4_path else '未指定'}' 未选择或不存在，跳过图片插入。",
                False)

        wb.save()
        wb.close()
        # FIXED: Always pass is_error
        if log_callback: log_callback("\n✅ 所有数据及图片已成功汇总到目标文件。", False)
        return True

    except Exception as e:
        # FIXED: Always pass is_error
        if log_callback: log_callback(f"❌ 出现错误：{e}", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback("请确保：", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback("1. Microsoft Excel 已安装并可正常运行。", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback("2. 所有源文件和目标报告文件在操作过程中是关闭状态。", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback("3. 文件路径正确无误，且您有读写权限。", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback(
            "4. 目标工作表名称与配置一致（特别是 '遗留缺陷列表', '产品需求列表', '验收测试用例', '设备外观图'）。", True)
        # FIXED: Always pass is_error
        if log_callback: log_callback(f"详细错误信息: {traceback.format_exc()}", True)
        return False
    finally:
        if app:
            app.quit()
            # FIXED: Always pass is_error
            if log_callback: log_callback("xlwings 应用程序已关闭。", False)
